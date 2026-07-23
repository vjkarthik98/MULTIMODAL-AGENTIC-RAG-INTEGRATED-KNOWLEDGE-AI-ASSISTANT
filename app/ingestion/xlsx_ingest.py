"""
XLSX/XLS ingestor — Phase 1 per-modality refactor.

XlsxIngestor.extract() → List[RawExtract]   (extraction only; no chunking)
ingest()               → List[IngestedDocument]  (backward-compat; full pipeline)
"""

from __future__ import annotations

import asyncio
import hashlib
import os
import re
import tempfile
import time
import uuid
from pathlib import Path
from typing import Any

from opentelemetry import trace
from opentelemetry.trace import Status, StatusCode
from prometheus_client import Counter, Histogram

from app.core.config import settings
from app.ingestion.base_ingest import BaseIngestor
from app.ingestion.schema import IngestedDocument, RawExtract, UniversalMetadata
from app.utils.logger import get_logger

logger = get_logger(__name__)
tracer = trace.get_tracer(__name__)

_ingest_duration = Histogram(
    "xlsx_ingest_duration_seconds",
    "XLSX ingestion duration",
    ["status"],
)
_ingest_errors = Counter(
    "xlsx_ingest_errors_total",
    "XLSX ingestion errors by type",
    ["error_type"],
)
_EXTRACTS_TOTAL = Counter("magik_xlsx_extracts_total", "Total extracts produced by xlsx ingestor")
_EXTRACT_ERRORS = Counter("magik_xlsx_extract_errors_total", "Errors in xlsx ingestor")

_semaphore = asyncio.Semaphore(5)

_UNIT_SCALE_PATTERNS: list[tuple[Any, str]] = [
    (re.compile(r'\bin billions?\b|\(\$\s*billions?\)', re.IGNORECASE), "billions"),
    (re.compile(r'\bin millions?\b|\(\$\s*millions?\)', re.IGNORECASE), "millions"),
    (
        re.compile(r'\bin thousands?\b|\busd ?000s?\b|\(\$\s*thousands?\)', re.IGNORECASE),
        "thousands",
    ),
]

_PNL_SHEET_KW: set = {"income", "p&l", "profit", "earnings", "loss", "pnl", "pl"}

_PNL_GROUP_MARKERS: list[tuple[Any, str]] = [
    (re.compile(r'\bnet income\b|\bnet profit\b|\bnet loss\b', re.IGNORECASE), "Per-Share Data"),
    (
        re.compile(r'\bebit\b|\boperating income\b|\boperating profit\b', re.IGNORECASE),
        "Below-the-Line",
    ),
    (re.compile(r'\bgross profit\b', re.IGNORECASE), "Cost Lines"),
    (re.compile(r'\brevenue\b|\bnet sales\b|\btotal revenue\b', re.IGNORECASE), "Revenue Lines"),
]

_ACCOUNTING_NEG_RE = re.compile(r'^\s*\([\d,]+\.?\d*\)\s*$')
_PERCENTAGE_RE = re.compile(r'^\s*-?\d+\.?\d*\s*%\s*$')
_MULTIPLE_RE = re.compile(r'^\s*-?\d+\.?\d*\s*[xX]\s*$')
_ACCOUNTING_FMT_RE = re.compile(r'_\)|#,##0.*\(', re.IGNORECASE)
_PERCENT_FMT_RE = re.compile(r'%')


def _format_percent_cell(value: float) -> str:
    """Render a percentage-formatted cell using its OWN display scale (x100),
    not the raw stored fraction. openpyxl returns the underlying float (e.g.
    0.0466 for a cell Excel displays as "4.66%"); passing that raw fraction
    into embeddings/context with no "%" marker is ambiguous, and the LLM has
    been observed reproducing it verbatim with a bare "%" appended — a 100x
    magnitude error (accuracy phase 2026-07)."""
    s = f"{value * 100:.3f}"
    if "." in s:
        int_part, dec_part = s.split(".", 1)
        dec_part = dec_part.rstrip("0")
        s = int_part + (f".{dec_part}" if dec_part else "")
    return s + "%"


# ─── Utilities ────────────────────────────────────────────────────────────────


def _classify_semantic_group(row_text: str) -> str:
    for pattern, group in _PNL_GROUP_MARKERS:
        if pattern.search(row_text):
            return group
    return ""


def _detect_display_format(value_str: str, number_format: str = "") -> str:
    """Classify a cell's display format from its string value or number_format attribute."""
    if _ACCOUNTING_NEG_RE.match(value_str):
        return "accounting_negative"
    if _PERCENTAGE_RE.match(value_str):
        return "percentage"
    if _MULTIPLE_RE.match(value_str):
        return "multiple"
    if number_format and _ACCOUNTING_FMT_RE.search(number_format):
        return "accounting_negative"
    return "standard"


def _extract_named_ranges(wb: Any) -> dict[str, Any]:
    """Extract workbook-level named ranges and resolve single-cell references to values."""
    result: dict[str, Any] = {}
    try:
        items: list[tuple[str, Any]] = []
        try:
            items = [(name, dn) for name, dn in wb.defined_names.items()]
        except AttributeError:
            try:
                items = [(dn.name, dn) for dn in wb.defined_names.definedName]
            except Exception:
                pass
        for name, dn in items:
            if name.startswith("_") or not name:
                continue
            try:
                for sheet_title, coord in dn.destinations:
                    try:
                        ws = wb[sheet_title]
                        clean_coord = coord.replace("$", "")
                        if ":" not in clean_coord:
                            val = ws[clean_coord].value
                            if val is not None:
                                result[name] = val
                        else:
                            result[name] = f"{sheet_title}!{coord}"
                    except Exception:
                        pass
            except Exception:
                pass
    except Exception:
        pass
    return result


def _sheet_type_classify(sheet_name: str) -> str:
    _MAP = {
        "income": "income_statement",
        "p&l": "income_statement",
        "profit": "income_statement",
        "earnings": "income_statement",
        "balance": "balance_sheet",
        "cash": "cash_flow",
        "capex": "capex",
        "assump": "assumptions",
        "input": "assumptions",
        "dcf": "model",
        "lbo": "model",
        "model": "model",
        "scenario": "scenarios",
        "case": "scenarios",
        "pivot": "pivot",
        "chart": "charts",
        "dashboard": "dashboard",
        "summary": "summary",
        "data": "data",
    }
    nl = sheet_name.lower()
    for key, stype in _MAP.items():
        if key in nl:
            return stype
    return "data"


def _load_all_rows_with_meta(
    ws: Any,
) -> tuple[list[dict], list[dict], str]:
    """Return (visible_row_metas, hidden_row_metas, unit_scale).

    Each row_meta: {"cells": List[str], "display_formats": List[str],
                    "is_hidden": bool, "row_idx": int}
    """
    from openpyxl.utils import get_column_letter

    hidden_col_letters = {col for col, cd in ws.column_dimensions.items() if cd.hidden}
    visible: list[dict] = []
    hidden_rows: list[dict] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        rd = ws.row_dimensions.get(row_idx)
        is_hidden = bool(rd and rd.hidden)
        cells: list[str] = []
        formats: list[str] = []
        for cell in row:
            col_letter = get_column_letter(cell.column)
            if col_letter in hidden_col_letters:
                continue
            val = cell.value
            number_format = getattr(cell, "number_format", "") or ""
            if (
                isinstance(val, (int, float))
                and not isinstance(val, bool)
                and _PERCENT_FMT_RE.search(number_format)
            ):
                val_str = _format_percent_cell(float(val))
            else:
                val_str = str(val if val is not None else "").strip()
            fmt = _detect_display_format(val_str, number_format)
            cells.append(val_str)
            formats.append(fmt)

        if not any(cells):
            continue

        meta = {
            "cells": cells,
            "display_formats": formats,
            "is_hidden": is_hidden,
            "row_idx": row_idx,
        }
        if is_hidden:
            hidden_rows.append(meta)
        else:
            visible.append(meta)

    flat_visible = [r["cells"] for r in visible]
    unit_scale = _detect_unit_scale(ws.title, flat_visible[:5]) if flat_visible else "units"
    return visible, hidden_rows, unit_scale


def _file_hash(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _quality(text: str) -> float:
    l = len(text)
    if l < 50:
        return 0.2
    if l < 200:
        return 0.5
    return 1.0


def _detect_unit_scale(sheet_name: str, rows: list[list[str]]) -> str:
    search_text = sheet_name + " " + " ".join(c for row in rows[:5] for c in row)
    for pattern, scale in _UNIT_SCALE_PATTERNS:
        if pattern.search(search_text):
            return scale
    return "units"


def _is_pnl_sheet(sheet_name: str) -> bool:
    return any(kw in sheet_name.lower() for kw in _PNL_SHEET_KW)


def _table_to_text(rows: Any) -> str:
    cleaned = [
        [str(cell or "").strip() for cell in row]
        for row in (rows or [])
        if any(cell for cell in row)
    ]
    if not cleaned:
        return ""
    return "\n".join(" | ".join(row) for row in cleaned)


def _pii_scrub(text: str, surface: str) -> str:
    try:
        from app.guardrails.pii import scrub_pii as _gp_scrub

        cleaned, _ = _gp_scrub(text)
        return cleaned
    except Exception:
        return text


def _sanitize_text(text: str, surface: str) -> str:
    try:
        from app.guardrails.input_guard import sanitize as _g

        return _g(text, surface=surface)
    except Exception:
        return text


def _try_float(v: Any) -> float | None:
    try:
        f = float(str(v).replace(",", ""))
        return f if str(f) != "nan" else None
    except (ValueError, TypeError):
        return None


def _load_worksheet_rows(ws: Any) -> tuple[list[list[str]], str]:
    """Load visible rows + detect unit scale. Returns (rows, unit_scale)."""
    from openpyxl.utils import get_column_letter

    hidden_col_letters = {
        col_letter for col_letter, cd in ws.column_dimensions.items() if cd.hidden
    }
    all_rows: list[list[str]] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        rd = ws.row_dimensions.get(row_idx)
        if rd and rd.hidden:
            continue
        cells = [
            str(c if c is not None else "").strip()
            for col_idx, c in enumerate(row, start=1)
            if get_column_letter(col_idx) not in hidden_col_letters
        ]
        all_rows.append(cells)
    non_empty = [r for r in all_rows if any(c for c in r)]
    unit_scale = _detect_unit_scale(ws.title, non_empty[:5]) if non_empty else "units"
    return non_empty, unit_scale


def _build_stats_summary(sheet_name: str, non_empty: list[list[str]]) -> str | None:
    """Compute a year-level statistics summary for time-series sheets."""
    if len(non_empty) <= 10:
        return None
    try:
        _num_cols = (
            [
                ci
                for ci, _ in enumerate(non_empty[0])
                if any(
                    _try_float(r[ci]) is not None for r in non_empty[1 : min(10, len(non_empty))]
                )
            ]
            if len(non_empty) > 1
            else []
        )

        _date_col = next(
            (
                ci
                for ci, _ in enumerate(non_empty[0])
                if any(
                    str(r[ci]).startswith("20") or str(r[ci]).startswith("19")
                    for r in non_empty[1 : min(5, len(non_empty))]
                    if r[ci]
                )
            ),
            None,
        )

        if not _num_cols or _date_col is None:
            return None

        _hdr = non_empty[0]
        _data_rows = [
            r
            for r in non_empty[1:]
            if r[_date_col] and any(_try_float(r[c]) is not None for c in _num_cols)
        ]
        if len(_data_rows) <= 10:
            return None

        _nc = _num_cols[0]
        _vals = [
            (r[_date_col], _try_float(r[_nc])) for r in _data_rows if _try_float(r[_nc]) is not None
        ]
        if not _vals:
            return None

        _max_v = max(_vals, key=lambda x: x[1])
        _min_v = min(_vals, key=lambda x: x[1])
        _first = _vals[0]
        _last = _vals[-1]

        _year_stats: dict = {}
        for _d, _v in _vals:
            _yr = str(_d)[:4]
            if _yr not in _year_stats:
                _year_stats[_yr] = {
                    "first": (_d, _v),
                    "last": (_d, _v),
                    "sum": 0.0,
                    "cnt": 0,
                    "min": (_d, _v),
                    "max": (_d, _v),
                }
            s = _year_stats[_yr]
            s["last"] = (_d, _v)
            s["sum"] += float(_v)
            s["cnt"] += 1
            if float(_v) < float(s["min"][1]):
                s["min"] = (_d, _v)
            if float(_v) > float(s["max"][1]):
                s["max"] = (_d, _v)

        _num_col_name = _hdr[_nc] if _hdr[_nc] else f"Column {_nc + 1}"
        _lines = [
            f"[COMPUTED SUMMARY — Sheet: {sheet_name}]",
            f"Column: {_num_col_name}",
            f"Dataset: {_first[0]} to {_last[0]} ({len(_vals)} trading days)",
            f"Overall maximum: {_max_v[1]:.2f} on {_max_v[0]}",
            f"Overall minimum: {_min_v[1]:.2f} on {_min_v[0]}",
            f"First value: {_first[1]:.2f} on {_first[0]}",
            f"Last value: {_last[1]:.2f} on {_last[0]}",
        ]
        for _yr, s in sorted(_year_stats.items()):
            _avg = s["sum"] / s["cnt"] if s["cnt"] else 0
            _pct = (
                (float(s["last"][1]) - float(s["first"][1])) / float(s["first"][1]) * 100
                if float(s["first"][1])
                else 0
            )
            _lines.append(
                f"Year {_yr}: open={float(s['first'][1]):.2f} ({s['first'][0]}), "
                f"close={float(s['last'][1]):.2f} ({s['last'][0]}), "
                f"avg={_avg:.2f}, high={float(s['max'][1]):.2f} ({s['max'][0]}), "
                f"low={float(s['min'][1]):.2f} ({s['min'][0]}), "
                f"change={_pct:+.2f}%, trading_days={s['cnt']}"
            )
        _sorted_yrs = sorted(_year_stats.keys())
        if len(_sorted_yrs) >= 2:
            for i in range(len(_sorted_yrs) - 1):
                _ya, _yb = _sorted_yrs[i], _sorted_yrs[i + 1]
                _va_open = float(_year_stats[_ya]["first"][1])
                _vb_close = float(_year_stats[_yb]["last"][1])
                if _va_open:
                    _cross_pct = (_vb_close - _va_open) / _va_open * 100
                    _lines.append(
                        f"Cross-period: start of {_ya} ({_year_stats[_ya]['first'][0]}, {_va_open:.2f}) "
                        f"to end of {_yb} ({_year_stats[_yb]['last'][0]}, {_vb_close:.2f}) "
                        f"= {_cross_pct:+.2f}%"
                    )
        return "\n".join(_lines)
    except Exception as exc:
        logger.warning("excel_stats_summary_failed", sheet=sheet_name, error=str(exc))
        return None


def _extract_chart_text(ws: Any, sheet_name: str) -> list[str]:
    chart_texts: list[str] = []
    try:
        for chart_idx, chart in enumerate(getattr(ws, "_charts", []) or []):
            try:

                def _safe_str(v: Any) -> str:
                    try:
                        return str(v) if v is not None else ""
                    except Exception:
                        return ""

                title_text = ""
                try:
                    t = chart.title
                    if t is not None:
                        if hasattr(t, "tx") and t.tx and getattr(t.tx, "rich", None):
                            runs = []
                            for p in t.tx.rich.p or []:
                                for r in p.r or []:
                                    runs.append(_safe_str(getattr(r, "t", "")))
                            title_text = " ".join(r for r in runs if r).strip()
                        elif hasattr(t, "tx") and t.tx and getattr(t.tx, "strRef", None):
                            title_text = _safe_str(getattr(t.tx.strRef, "f", ""))
                        else:
                            title_text = _safe_str(t)
                except Exception:
                    pass

                chart_type = type(chart).__name__
                x_axis = ""
                y_axis = ""
                try:
                    x_axis = _safe_str(getattr(getattr(chart, "x_axis", None), "title", "") or "")
                    y_axis = _safe_str(getattr(getattr(chart, "y_axis", None), "title", "") or "")
                except Exception:
                    pass

                series_names: list[str] = []
                data_refs: list[str] = []
                try:
                    for s in getattr(chart, "series", []) or []:
                        try:
                            if getattr(s, "tx", None) and getattr(s.tx, "strRef", None):
                                series_names.append(_safe_str(s.tx.strRef.f))
                        except Exception:
                            pass
                        try:
                            ref = getattr(getattr(s, "val", None), "numRef", None)
                            if ref is not None:
                                data_refs.append(_safe_str(getattr(ref, "f", "")))
                        except Exception:
                            pass
                except Exception:
                    pass

                parts: list[str] = [f"[Chart: {chart_type} on sheet {sheet_name}]"]
                if title_text:
                    parts.append(f"Title: {title_text}")
                if x_axis:
                    parts.append(f"X axis: {x_axis}")
                if y_axis:
                    parts.append(f"Y axis: {y_axis}")
                series_names = [s for s in series_names if s]
                if series_names:
                    parts.append("Series: " + ", ".join(series_names[:12]))
                data_refs = [r for r in data_refs if r]
                if data_refs:
                    parts.append("Data: " + "; ".join(data_refs[:6]))

                chart_text = "\n".join(parts)
                if len(chart_text) >= 16:
                    chart_texts.append(chart_text)
            except Exception as exc:
                logger.warning("excel_chart_failed", sheet=sheet_name, error=str(exc))
    except Exception as exc:
        logger.warning("excel_chart_scan_failed", sheet=sheet_name, error=str(exc))
    return chart_texts


# ─── Phase 1: XlsxIngestor ────────────────────────────────────────────────────


class XlsxIngestor(BaseIngestor):
    """Extracts raw content from XLSX/XLS files → List[RawExtract].

    Phase 1 responsibility: file I/O, sheet/row/chart parsing, security gates.
    Does NOT chunk. The chunker (Phase 2) handles splitting.
    """

    def health_check(self) -> dict:
        return {
            "modality": "xlsx",
            "status": "ok",
            "class": self.__class__.__name__,
        }

    async def extract(
        self,
        path: Path,
        metadata: UniversalMetadata,
    ) -> list[RawExtract]:
        source = path.name
        file_path = str(path)
        logger.info(
            event="extraction_start", modality="xlsx", file=str(path), size=path.stat().st_size
        )
        try:
            try:
                import openpyxl

                wb = openpyxl.load_workbook(file_path, data_only=True)
            except Exception as exc:
                raise ValueError(f"CORRUPTED_FILE: {exc}")

            extracts: list[RawExtract] = []

            try:
                # ── Named ranges (workbook-level) ─────────────────────────────
                named_ranges = _extract_named_ranges(wb)
                if named_ranges:
                    nr_lines = []
                    for nr_name, nr_val in named_ranges.items():
                        # Natural-language: "WACC (Weighted Average Cost of Capital) = 8.5%"
                        human = nr_name  # keep raw name; BM25 will tokenise it
                        nr_lines.append(f"{human} = {nr_val}")
                    nr_text = "\n".join(nr_lines)
                    nr_text = self._sanitize(nr_text, surface="excel_named_range_ingest")
                    if nr_text:
                        extracts.append(
                            RawExtract(
                                text=nr_text,
                                extract_type="named_range",
                                sheet=None,
                                raw_source_ref=f"xlsx:{path.name}|named_ranges",
                                extra={"named_ranges": named_ranges, "is_named_range_block": True},
                            )
                        )

                # ── Per-sheet extraction ──────────────────────────────────────
                for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                    try:
                        ws = wb[sheet_name]
                        visible_rows, hidden_rows_meta, unit_scale = _load_all_rows_with_meta(ws)
                        flat_visible = [r["cells"] for r in visible_rows]

                        # Cap rows per sheet to avoid runaway processing on huge spreadsheets.
                        max_rows = getattr(settings, "EXCEL_MAX_ROWS", 50_000)
                        if len(flat_visible) > max_rows:
                            logger.warning(
                                "excel_sheet_rows_capped",
                                sheet=sheet_name,
                                original=len(flat_visible),
                                capped=max_rows,
                                file=path.name,
                            )
                            flat_visible = flat_visible[:max_rows]
                            visible_rows = visible_rows[:max_rows]
                            hidden_rows_meta = [
                                r for r in hidden_rows_meta if r.get("row_idx", 0) <= max_rows
                            ]

                        if not flat_visible and not hidden_rows_meta:
                            logger.info(
                                "excel_empty_sheet_skipped", sheet=sheet_name, file=path.name
                            )
                            continue

                        sheet_type = _sheet_type_classify(sheet_name)
                        has_charts = bool(getattr(ws, "_charts", []))
                        has_pivot = bool(
                            getattr(ws, "_pivots", []) or getattr(ws, "PivotTableList", [])
                        )
                        dimensions = ws.dimensions or ""

                        # Sheet metadata extract
                        extracts.append(
                            RawExtract(
                                text="",
                                extract_type="sheet_metadata",
                                sheet=sheet_name,
                                raw_source_ref=f"xlsx:{path.name}|sheet:{sheet_name}|meta",
                                extra={
                                    "sheet_index": sheet_idx,
                                    "sheet_type": sheet_type,
                                    "has_charts": has_charts,
                                    "has_pivot": has_pivot,
                                    "dimensions": dimensions,
                                    "unit_scale": unit_scale,
                                },
                            )
                        )

                        # Unit header
                        if unit_scale != "units" and flat_visible:
                            unit_hdr = " ".join(flat_visible[0])
                            unit_hdr = self._sanitize(unit_hdr, surface="excel_ingest")
                            if unit_hdr:
                                extracts.append(
                                    RawExtract(
                                        text=unit_hdr,
                                        extract_type="unit_header",
                                        sheet=sheet_name,
                                        raw_source_ref=f"xlsx:{path.name}|sheet:{sheet_name}|unit_header",
                                        extra={"unit_scale": unit_scale},
                                    )
                                )

                        ROWS_PER_CHUNK = settings.EXCEL_ROWS_PER_CHUNK
                        header_row = flat_visible[0] if flat_visible else None

                        # ── Visible rows → row extracts ───────────────────────
                        for batch_start in range(0, len(flat_visible), ROWS_PER_CHUNK):
                            batch_cells = flat_visible[batch_start : batch_start + ROWS_PER_CHUNK]
                            batch_meta = visible_rows[batch_start : batch_start + ROWS_PER_CHUNK]
                            row_start = batch_start + 1
                            row_end = batch_start + len(batch_cells)

                            if batch_start > 0 and header_row and batch_cells[0] != header_row:
                                batch_cells = [header_row] + batch_cells

                            txt = _table_to_text(batch_cells)
                            if not txt.strip():
                                continue

                            sem_group = _classify_semantic_group(txt)
                            chunk_text = f"[Sheet: {sheet_name}, Rows {row_start}-{row_end}]\n{txt}"
                            chunk_text = self._sanitize(chunk_text, surface="excel_ingest")
                            if not chunk_text.strip():
                                continue
                            chunk_text = self._scrub_pii(chunk_text, surface="excel_ingest")

                            # Dominant display format across batch
                            all_fmts = [f for r in batch_meta for f in r.get("display_formats", [])]
                            non_std = [f for f in all_fmts if f != "standard"]
                            dominant_fmt = (
                                max(set(non_std), key=non_std.count) if non_std else "standard"
                            )

                            extracts.append(
                                RawExtract(
                                    text=chunk_text,
                                    extract_type="table_row",
                                    sheet=sheet_name,
                                    raw_source_ref=f"xlsx:{path.name}|sheet:{sheet_name}|rows:{row_start}-{row_end}",
                                    extra={
                                        "unit_scale": unit_scale,
                                        "row_start": row_start,
                                        "row_end": row_end,
                                        "row_num": row_start,
                                        "semantic_group": sem_group,
                                        "display_format": dominant_fmt,
                                        "sheet_index": sheet_idx,
                                        "sheet_type": sheet_type,
                                    },
                                )
                            )

                        # ── Hidden rows → table_row_hidden extracts ───────────
                        if hidden_rows_meta:
                            for h_row in hidden_rows_meta:
                                h_txt = " | ".join(c for c in h_row["cells"] if c)
                                if not h_txt:
                                    continue
                                h_txt = self._sanitize(h_txt, surface="excel_ingest")
                                h_txt = self._scrub_pii(h_txt, surface="excel_ingest")
                                if h_txt:
                                    extracts.append(
                                        RawExtract(
                                            text=h_txt,
                                            extract_type="table_row_hidden",
                                            sheet=sheet_name,
                                            raw_source_ref=f"xlsx:{path.name}|sheet:{sheet_name}|hidden_row:{h_row['row_idx']}",
                                            extra={
                                                "unit_scale": unit_scale,
                                                "row_num": h_row["row_idx"],
                                                "is_hidden": True,
                                                "sheet_index": sheet_idx,
                                                "sheet_type": sheet_type,
                                            },
                                        )
                                    )

                        # ── Statistics summary ────────────────────────────────
                        summary = _build_stats_summary(sheet_name, flat_visible)
                        if summary:
                            extracts.append(
                                RawExtract(
                                    text=summary,
                                    extract_type="named_range",
                                    sheet=sheet_name,
                                    raw_source_ref=f"xlsx:{path.name}|sheet:{sheet_name}|stats_summary",
                                    extra={"is_stats_summary": True, "unit_scale": unit_scale},
                                )
                            )

                        # ── Charts ────────────────────────────────────────────
                        for chart_text in _extract_chart_text(ws, sheet_name):
                            chart_text = self._scrub_pii(chart_text, surface="excel_chart_ingest")
                            if chart_text:
                                extracts.append(
                                    RawExtract(
                                        text=chart_text,
                                        extract_type="chart_image",
                                        sheet=sheet_name,
                                        raw_source_ref=f"xlsx:{path.name}|sheet:{sheet_name}|chart",
                                    )
                                )

                        # ── Embedded images ───────────────────────────────────
                        try:
                            for img_idx, img_obj in enumerate(getattr(ws, "_images", []) or []):
                                try:
                                    blob: bytes | None = None
                                    ext_hint = ".png"
                                    data_attr = getattr(img_obj, "_data", None)
                                    if callable(data_attr):
                                        try:
                                            blob = data_attr()
                                        except Exception:
                                            blob = None
                                    if not blob:
                                        ref = getattr(img_obj, "ref", None) or getattr(
                                            img_obj, "path", None
                                        )
                                        if ref and hasattr(ref, "read"):
                                            try:
                                                ref.seek(0)
                                            except Exception:
                                                pass
                                            blob = ref.read()
                                        elif isinstance(ref, (str, os.PathLike)):
                                            try:
                                                with open(ref, "rb") as fh:
                                                    blob = fh.read()
                                                ext_hint = os.path.splitext(str(ref))[1] or ext_hint
                                            except Exception:
                                                blob = None
                                    fmt = getattr(img_obj, "format", None)
                                    if fmt:
                                        ext_hint = "." + str(fmt).lower()
                                    if blob and len(blob) >= 256:
                                        extracts.append(
                                            RawExtract(
                                                text="",
                                                extract_type="image_region",
                                                sheet=sheet_name,
                                                raw_source_ref=f"xlsx:{path.name}|sheet:{sheet_name}|img:{img_idx}",
                                                raw_bytes=blob,
                                                extra={"img_ext": ext_hint},
                                            )
                                        )
                                except Exception as exc:
                                    logger.warning(
                                        "excel_embedded_image_failed",
                                        sheet=sheet_name,
                                        error=str(exc),
                                    )
                        except Exception as exc:
                            logger.warning(
                                "excel_image_scan_failed", sheet=sheet_name, error=str(exc)
                            )

                    except ValueError:
                        raise
                    except Exception as exc:
                        logger.warning("excel_sheet_failed", sheet=sheet_name, error=str(exc))
            finally:
                try:
                    wb.close()
                except Exception:
                    pass

            if not extracts:
                raise ValueError("NO_EXTRACTS_PRODUCED")

            _EXTRACTS_TOTAL.inc(len(extracts))
            logger.info(
                event="extraction_complete", modality="xlsx", file=str(path), extracts=len(extracts)
            )
            return extracts
        except Exception as _exc:
            _EXTRACT_ERRORS.inc()
            logger.error(event="extraction_failed", modality="xlsx", source=source, error=str(_exc))
            raise


# ─── Backward-compat ingest() — full pipeline ─────────────────────────────────


def _base_structure(doc_id: str, session_id: str, source_path: str, **extra: Any) -> dict[str, Any]:
    return {"doc_id": doc_id, "session_id": session_id, "source_path": source_path, **extra}


def _ingest_embedded_image_bytes(
    blob: bytes,
    ext_hint: str,
    session_id: str,
    parent_doc_id: str,
    parent_modality: str,
    parent_source: str,
    parent_page: int | None = None,
    parent_sheet: str | None = None,
) -> list[IngestedDocument]:
    if not blob or len(blob) < 256:
        return []
    safe_ext = (ext_hint or ".png").lower().strip()
    if safe_ext == ".jpeg":
        safe_ext = ".jpg"
    if safe_ext not in {".png", ".jpg", ".bmp", ".gif", ".webp", ".tiff", ".tif"}:
        safe_ext = ".png"
    from app.utils.paths import resolved_temp_dir

    tmp_dir = resolved_temp_dir() / "embedded_images"
    try:
        tmp_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        tmp_dir = Path(tempfile.gettempdir())
    digest = hashlib.sha256(blob).hexdigest()
    tmp_path = tmp_dir / f"{digest}{safe_ext}"
    try:
        if not tmp_path.exists():
            tmp_path.write_bytes(blob)
        from app.ingestion.image_ingest import ingest as image_ingest

        return (
            image_ingest(
                str(tmp_path),
                session_id,
                parent_doc_id=parent_doc_id,
                parent_modality=parent_modality,
                parent_source=parent_source,
                parent_page=parent_page,
                parent_sheet=parent_sheet,
            )
            or []
        )
    except Exception as exc:
        logger.warning(
            "embedded_image_ingest_failed",
            parent_source=parent_source,
            sheet=parent_sheet,
            page=parent_page,
            error=str(exc),
        )
        return []


async def ingest(file_path: str, session_id: str) -> list[IngestedDocument]:
    """Backward-compatible entry point. Router imports this until Phase 8."""
    if not session_id:
        raise ValueError("SESSION_ID_REQUIRED")

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"FILE_NOT_FOUND: {file_path}")

    file_size = path.stat().st_size
    if file_size == 0:
        raise ValueError("EMPTY_FILE")
    if file_size > settings.MAX_FILE_SIZE_XLSX:
        raise ValueError(f"FILE_TOO_LARGE: {file_size}")

    with tracer.start_as_current_span("xlsx_ingest") as span:
        span.set_attribute("file.name", path.name)
        span.set_attribute("file.size", file_size)
        span.set_attribute("session.id", session_id)
        start = time.time()

        async with _semaphore:
            try:
                logger.info(
                    "xlsx_ingest_start", file=path.name, size=file_size, session_id=session_id
                )

                doc_id = str(uuid.uuid4())
                source_name = path.name
                source_path_str = str(path.resolve())
                ROWS_PER_CHUNK = settings.EXCEL_ROWS_PER_CHUNK

                try:
                    import openpyxl

                    wb = openpyxl.load_workbook(file_path, data_only=True)
                except Exception as exc:
                    raise ValueError(f"CORRUPTED_FILE: {exc}")

                documents: list[IngestedDocument] = []
                doc_warnings: list[str] = []

                try:
                    for sheet_name in wb.sheetnames:
                        try:
                            ws = wb[sheet_name]
                            non_empty, unit_scale = _load_worksheet_rows(ws)

                            if not non_empty:
                                doc_warnings.append(f"Empty sheet skipped: {sheet_name}")
                                continue

                            header_row = non_empty[0] if non_empty else None

                            for batch_start in range(0, len(non_empty), ROWS_PER_CHUNK):
                                batch = non_empty[batch_start : batch_start + ROWS_PER_CHUNK]
                                row_start = batch_start + 1
                                row_end = batch_start + len(batch)

                                if batch_start > 0 and header_row and batch[0] != header_row:
                                    batch = [header_row] + batch

                                txt = _table_to_text(batch)
                                if not txt.strip():
                                    continue

                                chunk_text = (
                                    f"[Sheet: {sheet_name}, Rows {row_start}-{row_end}]\n{txt}"
                                )
                                try:
                                    from app.guardrails.input_guard import (
                                        sanitize as _guard_sanitize,
                                    )

                                    _clean = _guard_sanitize(chunk_text, surface="excel_ingest")
                                    if _clean != chunk_text:
                                        logger.warning(
                                            "excel_injection_sanitized",
                                            file=source_name,
                                            sheet=sheet_name,
                                        )
                                        chunk_text = _clean
                                except Exception as _ge:
                                    logger.warning(
                                        "excel_guardrail_failed", file=source_name, error=str(_ge)
                                    )

                                if not chunk_text.strip():
                                    continue
                                chunk_text = _pii_scrub(chunk_text, surface="excel_ingest")

                                documents.append(
                                    IngestedDocument(
                                        text=chunk_text,
                                        modality="table",
                                        subtype="structured",
                                        source_type="excel",
                                        source=source_name,
                                        structure=_base_structure(
                                            doc_id,
                                            session_id,
                                            source_path_str,
                                            sheet=sheet_name,
                                            row_start=row_start,
                                            row_end=row_end,
                                            page_number=None,
                                            total_pages=None,
                                            section_title=sheet_name,
                                            ingestion_timestamp=time.time(),
                                            language="en",
                                            file_size_bytes=file_size,
                                            content_type="excel_sheet",
                                            ingestion_time=time.time(),
                                        ),
                                        extra_metadata={
                                            "data_quality_score": _quality(txt),
                                            "importance_score": _quality(txt),
                                            "modality_weight": 1.0,
                                        },
                                    ).finalize()
                                )

                            # Stats summary
                            summary = _build_stats_summary(sheet_name, non_empty)
                            if summary:
                                documents.append(
                                    IngestedDocument(
                                        text=summary,
                                        modality="table",
                                        subtype="summary",
                                        source_type="excel",
                                        source=source_name,
                                        structure=_base_structure(
                                            doc_id,
                                            session_id,
                                            source_path_str,
                                            sheet=sheet_name,
                                            row_start=1,
                                            row_end=len(non_empty),
                                            page_number=None,
                                            total_pages=None,
                                            section_title=sheet_name,
                                            ingestion_timestamp=time.time(),
                                            language="en",
                                            file_size_bytes=file_size,
                                            content_type="excel_summary",
                                            ingestion_time=time.time(),
                                        ),
                                        extra_metadata={
                                            "data_quality_score": 1.0,
                                            "importance_score": 1.0,
                                            "modality_weight": 1.5,
                                        },
                                    ).finalize()
                                )

                            # Embedded images
                            try:
                                for img_obj in getattr(ws, "_images", []) or []:
                                    try:
                                        blob: bytes | None = None
                                        ext_hint = ".png"
                                        data_attr = getattr(img_obj, "_data", None)
                                        if callable(data_attr):
                                            try:
                                                blob = data_attr()
                                            except Exception:
                                                blob = None
                                        if not blob:
                                            ref = getattr(img_obj, "ref", None) or getattr(
                                                img_obj, "path", None
                                            )
                                            if ref and hasattr(ref, "read"):
                                                try:
                                                    ref.seek(0)
                                                except Exception:
                                                    pass
                                                blob = ref.read()
                                            elif isinstance(ref, (str, os.PathLike)):
                                                try:
                                                    with open(ref, "rb") as fh:
                                                        blob = fh.read()
                                                    ext_hint = (
                                                        os.path.splitext(str(ref))[1] or ext_hint
                                                    )
                                                except Exception:
                                                    blob = None
                                        fmt = getattr(img_obj, "format", None)
                                        if fmt:
                                            ext_hint = "." + str(fmt).lower()
                                        embedded = _ingest_embedded_image_bytes(
                                            blob=blob or b"",
                                            ext_hint=ext_hint,
                                            session_id=session_id,
                                            parent_doc_id=doc_id,
                                            parent_modality="excel",
                                            parent_source=source_name,
                                            parent_sheet=sheet_name,
                                        )
                                        documents.extend(embedded)
                                    except Exception as exc:
                                        logger.warning(
                                            "excel_embedded_image_failed",
                                            sheet=sheet_name,
                                            error=str(exc),
                                        )
                            except Exception as exc:
                                logger.warning(
                                    "excel_image_scan_failed", sheet=sheet_name, error=str(exc)
                                )

                            # Charts
                            for chart_text in _extract_chart_text(ws, sheet_name):
                                chart_text = _pii_scrub(chart_text, surface="excel_chart_ingest")
                                if len(chart_text) < 16:
                                    continue
                                documents.append(
                                    IngestedDocument(
                                        text=chart_text,
                                        modality="table",
                                        subtype="chart",
                                        source_type="excel",
                                        source=source_name,
                                        structure=_base_structure(
                                            doc_id,
                                            session_id,
                                            source_path_str,
                                            sheet=sheet_name,
                                            page_number=None,
                                            total_pages=None,
                                            section_title=None,
                                            ingestion_timestamp=time.time(),
                                            language="en",
                                            file_size_bytes=file_size,
                                            content_type="excel_chart",
                                            ingestion_time=time.time(),
                                        ),
                                        extra_metadata={
                                            "data_quality_score": _quality(chart_text),
                                            "importance_score": 0.9,
                                            "modality_weight": 1.0,
                                        },
                                    ).finalize()
                                )

                        except ValueError:
                            raise
                        except Exception as exc:
                            logger.warning("excel_sheet_failed", sheet=sheet_name, error=str(exc))
                finally:
                    try:
                        wb.close()
                    except Exception:
                        pass

                if not documents:
                    raise ValueError("NO_CONTENT_EXTRACTED")

                latency = round(time.time() - start, 2)
                _ingest_duration.labels(status="success").observe(latency)
                span.set_attribute("docs.count", len(documents))
                span.set_status(Status(StatusCode.OK))
                logger.info(
                    "xlsx_ingest_success",
                    file=path.name,
                    docs=len(documents),
                    latency=latency,
                    session_id=session_id,
                    warnings=doc_warnings or None,
                )
                return documents

            except Exception as exc:
                latency = round(time.time() - start, 2)
                error_type = type(exc).__name__
                _ingest_duration.labels(status="error").observe(latency)
                _ingest_errors.labels(error_type=error_type).inc()
                span.set_status(Status(StatusCode.ERROR, str(exc)))
                span.record_exception(exc)
                logger.error(
                    "xlsx_ingest_failed",
                    file=path.name,
                    session_id=session_id,
                    error=str(exc),
                    error_type=error_type,
                    latency=latency,
                )
                raise


def ingest_sync(file_path: str, session_id: str) -> list[IngestedDocument]:
    try:
        loop = asyncio.get_event_loop()
        if loop.is_running():
            import concurrent.futures

            with concurrent.futures.ThreadPoolExecutor() as pool:
                future = pool.submit(asyncio.run, ingest(file_path, session_id))
                return future.result()
        return loop.run_until_complete(ingest(file_path, session_id))
    except RuntimeError:
        return asyncio.run(ingest(file_path, session_id))


async def ingest_xlsx_full(file_path: str, session_id: str) -> list[IngestedDocument]:
    """Production path: XlsxIngestor → XlsxChunker with full Phase 1.4/2.4/3.4 metadata."""
    from app.chunking import chunk_raw_extracts
    from app.ingestion.schema import UniversalMetadata

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"FILE_NOT_FOUND: {file_path}")
    file_size = path.stat().st_size
    if file_size == 0:
        raise ValueError("EMPTY_FILE")
    if file_size > settings.MAX_FILE_SIZE_XLSX:
        raise ValueError(f"FILE_TOO_LARGE: {file_size}")

    meta = UniversalMetadata(
        source_path=str(path.resolve()),
        modality="xlsx",
        file_size_bytes=file_size,
        custom_fields={"session_id": session_id},
    )
    ingestor = XlsxIngestor()
    extracts = await ingestor.extract(path, meta)
    docs = chunk_raw_extracts(extracts, meta, "xlsx")
    for doc in docs:
        struct = getattr(doc, "structure", None)
        if struct is not None and struct.get("session_id") in (None, "default"):
            struct["session_id"] = session_id
    return docs
